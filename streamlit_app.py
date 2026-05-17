
import io
import math
from functools import lru_cache
from random import sample

import streamlit as st
import folium
from streamlit_folium import st_folium
import pandas as pd
import numpy as np
import requests

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from backend import (
    load_graph, load_club_traces, route, route_waypoints,
    similarity_to_club, f1_geo,
    export_gpx_string, read_gpx_string,
)


# ============================================================================
# CONFIG
# ============================================================================

st.set_page_config(
    page_title="Vélo Plaisir IdF",
    page_icon="🚴",
    layout="wide",
)

st.title("🚴 Planificateur d'itinéraires vélo — Île-de-France")
st.caption("Routes plaisir")

# ============================================================================
# SESSION STATE
# ============================================================================

def init_session_state():

    defaults = {
        "waypoints": [],
        "route_result": None,
        "uploaded_trace": None,
        "uploaded_elevs": None,
        "uploaded_sim": None,
        "comparison": None,
        "bumps_route": None,
        "bumps_uploaded": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v



@st.cache_resource(show_spinner="Chargement du graphe routier...")
def cached_graph():
    return load_graph()


@st.cache_resource(show_spinner="Chargement des traces club...")
def cached_traces():
    return load_club_traces()


@st.cache_data
def cached_simplified_traces(_traces):

    simplified = []
    for t in _traces:
        coords = t["coords"]
        n = len(coords)
        if n <= 100:
            simp = coords
        else:
            step = max(1, n // 100)
            simp = coords[::step]
            if simp[-1] != coords[-1]:
                simp.append(coords[-1])
        simplified.append({"name": t["name"], "coords": simp})
    return simplified



# ============================================================================
# GÉOCODAGE
# ============================================================================

@lru_cache(maxsize=200)
def geocode_address(address):

    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {
            "q": address,
            "format": "json",
            "limit": 1,
            "countrycodes": "fr",
            "viewbox": "1.2,49.3,3.6,48.0",
            "bounded": 1,
        }
        headers = {"User-Agent": "VeloClubIDF/1.0"}
        resp = requests.get(url, params=params, headers=headers, timeout=5)
        results = resp.json()
        if results:
            return float(results[0]["lat"]), float(results[0]["lon"]), results[0]["display_name"]
    except Exception as e:
        print(f"Geocoding error: {e}")
    return None


# ============================================================================
# DÉTECTION DES BOSSES
# ============================================================================

def haversine_m(p1, p2):
    R = 6371000
    phi1, phi2 = math.radians(p1[0]), math.radians(p2[0])
    dphi = math.radians(p2[0] - p1[0])
    dlam = math.radians(p2[1] - p1[1])
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1-a))


def _resample_elevation(coords, elevations, step_m=25):

    if len(coords) < 2 or len(elevations) != len(coords):
        return [], []
    cum = [0.0]
    for i in range(1, len(coords)):
        cum.append(cum[-1] + haversine_m(coords[i-1], coords[i]))
    total = cum[-1]
    dists, elevs = [], []
    d = 0.0
    j = 0
    while d <= total:
        while j < len(cum) - 1 and cum[j+1] < d:
            j += 1
        if j >= len(cum) - 1:
            e = elevations[-1]
        else:
            denom = max(cum[j+1] - cum[j], 0.001)
            frac = (d - cum[j]) / denom
            e = elevations[j] + frac * (elevations[j+1] - elevations[j])
        dists.append(d)
        elevs.append(e)
        d += step_m
    return dists, elevs


def _smooth(vals, window):
    out = []
    half = window // 2
    n = len(vals)
    for i in range(n):
        lo = max(0, i - half)
        hi = min(n, i + half + 1)
        out.append(sum(vals[lo:hi]) / (hi - lo))
    return out


def _solve_climb_speed(power_w, grade_pct, mass_kg=77, CdA=0.30, Crr=0.004):
    """Résout P = (m·g·(grade+Crr))·v + 0.5·rho·CdA·v^3 par Newton."""
    g, rho = 9.81, 1.225
    grade = grade_pct / 100
    F_lin = mass_kg * g * (grade + Crr)
    F_aero = 0.5 * rho * CdA
    v = 5.0
    for _ in range(50):
        f = F_lin * v + F_aero * v**3 - power_w
        df = F_lin + 3 * F_aero * v**2
        v_new = v - f / df
        if abs(v_new - v) < 0.0001:
            v = v_new
            break
        v = max(0.5, v_new)
    return v


def _format_time(sec):
    m = int(sec // 60)
    s = int(sec % 60)
    return f"{m}'{s:02d}" + '"'


def _find_pointu_section(smooth, dists, start_idx, peak_idx, bump_grade,
                          min_grade=6.5, min_delta=1.5, min_length=150):
    """Trouve la section avec la pente la plus forte qui dépasse les seuils.

    Critères stricts :
    - Pente >= 6.5%
    - Pente - pente_bosse >= 1.5 point (vraiment plus pentue que le reste)
    - Longueur >= 150m
    """
    step = dists[1] - dists[0] if len(dists) > 1 else 25
    bump_length = dists[peak_idx] - dists[start_idx]
    max_window = min(bump_length * 0.85, 500)

    best = None
    for window_m in range(int(min_length), int(max_window) + 25, 25):
        ws = max(2, int(window_m / step))
        if ws > peak_idx - start_idx:
            continue

        # Trouve la position avec la pente max pour cette fenêtre
        best_pos_grade = -1
        best_pos_i = None
        for i in range(start_idx, peak_idx - ws + 1):
            j = i + ws
            length = dists[j] - dists[i]
            gain = smooth[j] - smooth[i]
            if gain <= 0:
                continue
            grade = gain / length * 100
            if grade > best_pos_grade:
                best_pos_grade = grade
                best_pos_i = i
                best_pos_length = length

        if best_pos_i is None or best_pos_grade < min_grade:
            continue
        if best_pos_grade - bump_grade < min_delta:
            continue

        # Garde la plus pentue, à pente similaire préfère plus longue
        if best is None or best_pos_grade > best['grade'] + 0.3:
            best = {'start_km': dists[best_pos_i]/1000,
                    'length_m': best_pos_length,
                    'grade': best_pos_grade}
        elif abs(best_pos_grade - best['grade']) < 0.3 and best_pos_length > best['length_m']:
            best = {'start_km': dists[best_pos_i]/1000,
                    'length_m': best_pos_length,
                    'grade': best_pos_grade}

    if best:
        return f"{best['length_m']:.0f}m à {best['grade']:.1f}% (km {best['start_km']:.1f})"
    return "Non"


def _detect_bumps_one_pass(dists, smooth, mass_kg, CdA, Crr,
                            neg_grade_break=-2.0, drop_ratio=0.10):
    """Une passe de détection avec un smoothing donné."""
    n = len(smooth)
    step = 25
    window_samples = max(2, int(100 / step))

    local_grade = []
    for i in range(n):
        j_lo = max(0, i - window_samples//2)
        j_hi = min(n-1, i + window_samples//2)
        if dists[j_hi] - dists[j_lo] < 1:
            local_grade.append(0)
        else:
            g = (smooth[j_hi] - smooth[j_lo]) / (dists[j_hi] - dists[j_lo]) * 100
            local_grade.append(g)

    bumps_raw = []
    i = 0
    while i < n - 1:
        while i < n - 1 and local_grade[i] < 3.0:
            i += 1
        if i >= n - 1:
            break

        start_i = i
        peak_i = i
        peak_val = smooth[i]
        j = i + 1

        while j < n:
            if smooth[j] >= peak_val:
                peak_val = smooth[j]
                peak_i = j
                j += 1
            else:
                drop = peak_val - smooth[j]
                current_gain = peak_val - smooth[start_i]
                max_drop_tolerance = max(3.0, min(25.0, current_gain * drop_ratio))
                if drop > max_drop_tolerance:
                    break
                if local_grade[j] < neg_grade_break and j - start_i > 8:
                    break
                j += 1

        length = dists[peak_i] - dists[start_i]
        gain = smooth[peak_i] - smooth[start_i]

        if length >= 250 and gain > 0:
            grade = gain / length * 100
            if length >= 400:
                req = 3.0
            else:
                req = 7.0 - 4.0 * (length - 250) / 150
            if grade >= req:
                bumps_raw.append({
                    'start_i': start_i,
                    'peak_i': peak_i,
                    'km_start': dists[start_i] / 1000,
                    'km_end': dists[peak_i] / 1000,
                    'length_m': int(length),
                    'grade': round(grade, 1),
                    'gain': int(gain),
                })
        i = peak_i + 1

    return bumps_raw, local_grade, window_samples


def detect_bumps(coords, elevations,
                 mass_kg=77, CdA=0.30, Crr=0.004):
    """Détection des bosses en 2 passes :
    - Passe 1 (smooth=2) : capture les petites bosses (~300m)
    - Passe 2 (smooth=6) : capture les longs cols qui pourraient être fragmentés
    - Fusion : si un long col contient plusieurs petites bosses, on garde le col
    """
    if len(coords) < 4 or len(elevations) != len(coords):
        return []

    dists, elevs = _resample_elevation(coords, elevations, step_m=25)
    if len(dists) < 10:
        return []

    # Passe 1 : smooth léger pour petites bosses (~300m)
    smooth_fine = _smooth(elevs, 2)
    bumps_fine, _, _ = _detect_bumps_one_pass(
        dists, smooth_fine, mass_kg, CdA, Crr,
        neg_grade_break=-2.0, drop_ratio=0.10
    )

    # Passe 2 : smooth moyen pour cols avec petits replats (~3-5km)
    smooth_coarse = _smooth(elevs, 6)
    bumps_coarse, _, _ = _detect_bumps_one_pass(
        dists, smooth_coarse, mass_kg, CdA, Crr,
        neg_grade_break=-5.0, drop_ratio=0.10
    )

    # Passe 3 : smooth très fort pour grands cols avec replats marqués (lacets)
    smooth_macro = _smooth(elevs, 16)
    bumps_macro, _, _ = _detect_bumps_one_pass(
        dists, smooth_macro, mass_kg, CdA, Crr,
        neg_grade_break=-8.0, drop_ratio=0.20
    )

    # Smooth très fort pour calcul des sections pointues sur longs cols
    # Adaptatif : plus la bosse est longue, plus on lisse fort
    smooth_sections_strong = _smooth(elevs, 20)  # très fort pour cols > 5km
    smooth_sections = _smooth(elevs, 12)         # fort pour cols 1-5km

    # === FUSION EN 3 NIVEAUX ===
    # Priorité : macro > coarse > fine
    # Une bosse macro absorbe toutes les coarse et fine dedans
    # Une bosse coarse absorbe toutes les fine dedans (si >= 2 OU coarse 1.5x plus long)

    final_bumps = []

    # === FUSION PAR CHEVAUCHEMENT ===
    # Stratégie simple : on collecte toutes les bosses des 3 passes, on les trie
    # par longueur décroissante, et on garde une bosse seulement si elle ne
    # chevauche pas trop avec une bosse déjà gardée (plus longue).

    # Marque le niveau de chaque bosse pour distinguer macro/coarse/fine
    all_candidates = []
    for b in bumps_macro:
        all_candidates.append({**b, '_level': 'macro', 'is_coarse': True})
    for b in bumps_coarse:
        all_candidates.append({**b, '_level': 'coarse', 'is_coarse': True})
    for b in bumps_fine:
        all_candidates.append({**b, '_level': 'fine', 'is_coarse': False})

    # Trie par longueur DÉCROISSANTE : les longues bosses (macro) prioritaires
    all_candidates.sort(key=lambda b: -b['length_m'])

    final_bumps = []
    for cand in all_candidates:
        # Vérifie si cand chevauche significativement avec une bosse déjà acceptée
        overlap_too_much = False
        for kept in final_bumps:
            # Calcul du chevauchement
            overlap_start = max(cand['km_start'], kept['km_start'])
            overlap_end = min(cand['km_end'], kept['km_end'])
            overlap_km = max(0, overlap_end - overlap_start)
            cand_length_km = cand['length_m'] / 1000

            # Si plus de 50% de la bosse candidate chevauche une bosse gardée, on rejette
            if cand_length_km > 0 and overlap_km / cand_length_km > 0.5:
                overlap_too_much = True
                break

        if not overlap_too_much:
            final_bumps.append(cand)

    # Re-trie par km_start pour l'affichage
    final_bumps.sort(key=lambda b: b['km_start'])

    # Construit le résultat final avec tous les champs
    # On utilise smooth_fine pour calculer les détails (sections pointues etc.)
    # car c'est plus précis
    step = 25
    window_samples = max(2, int(100 / step))
    n = len(smooth_fine)

    result = []
    for b in final_bumps:
        # Recalcule les indices selon le smooth utilisé
        # (les bosses coarse pointent vers smooth_coarse, fines vers smooth_fine)
        smooth_used = smooth_coarse if b['is_coarse'] else smooth_fine
        start_i = b['start_i']
        peak_i = b['peak_i']
        length = dists[peak_i] - dists[start_i]
        gain = smooth_used[peak_i] - smooth_used[start_i]
        grade = gain / length * 100 if length > 0 else 0

        # Pente max : smoothing adaptatif selon la longueur de la bosse
        if not b['is_coarse']:
            smooth_for_max = smooth_fine
        elif (dists[b['peak_i']] - dists[b['start_i']]) > 5000:
            # Col > 5km : smoothing très fort
            smooth_for_max = smooth_sections_strong
        else:
            smooth_for_max = smooth_sections
        max_g = 0
        for k in range(start_i, peak_i - window_samples + 1):
            seg_g = (smooth_for_max[k + window_samples] - smooth_for_max[k]) / 100 * 100
            if seg_g > max_g:
                max_g = seg_g

        # Section pointue : même logique adaptative
        if not b['is_coarse']:
            smooth_for_pointu = smooth_fine
        elif (dists[b['peak_i']] - dists[b['start_i']]) > 5000:
            smooth_for_pointu = smooth_sections_strong
        else:
            smooth_for_pointu = smooth_sections
        section_pointue = _find_pointu_section(smooth_for_pointu, dists, start_i, peak_i, grade)

        # Analyse 2,5km après sommet sur le bon smooth
        idx_after = min(peak_i + int(2500 / 25), n - 1)
        if idx_after > peak_i:
            diff = smooth_used[idx_after] - smooth_used[peak_i]
            delta_d = dists[idx_after] - dists[peak_i]
            g_after = diff / delta_d * 100 if delta_d > 0 else 0
            # Seuils plus sensibles : ±0.5% suffit pour parler de descente/montée
            if g_after < -0.5:
                after = f"Descente, {diff:+.0f}m sur 2,5km"
            elif g_after > 0.5:
                after = "Montée continue"
            else:
                after = "Plat"
        else:
            after = "Fin du parcours"

        v250 = _solve_climb_speed(250, grade, mass_kg, CdA, Crr)
        v300 = _solve_climb_speed(300, grade, mass_kg, CdA, Crr)
        v350 = _solve_climb_speed(350, grade, mass_kg, CdA, Crr)

        fiets = gain**2 / (length * 10)
        if smooth_used[peak_i] > 1000:
            fiets += (smooth_used[peak_i] - 1000) / 1000

        if fiets >= 6.5: fiets_cat = "HC"
        elif fiets >= 5.0: fiets_cat = "Cat 1"
        elif fiets >= 3.5: fiets_cat = "Cat 2"
        elif fiets >= 2.0: fiets_cat = "Cat 3"
        elif fiets >= 0.5: fiets_cat = "Cat 4"
        elif fiets >= 0.25: fiets_cat = "Cat 5"
        else: fiets_cat = "—"

        result.append({
            "num": len(result) + 1,
            "km_start": dists[start_i] / 1000,
            "length_m": int(length),
            "grade_avg": round(grade, 1),
            "grade_max_100m": round(max_g, 1),
            "section_pointue": section_pointue,
            "after_summit": after,
            "time_250w": _format_time(length / v250),
            "time_300w": _format_time(length / v300),
            "time_350w": _format_time(length / v350),
            "d_plus_m": int(gain),
            "alt_summit_m": int(smooth_used[peak_i]),
            "fiets": round(fiets, 2),
            "fiets_cat": fiets_cat,
        })
    # Fusion finale : élimine les bosses qui se chevauchent
    result = _merge_overlapping_bumps(result)
    return result

def _merge_overlapping_bumps(bumps, overlap_threshold=0.3):
    """Fusionne les bosses qui se chevauchent.
    Si deux bosses partagent plus de overlap_threshold (30%) de la plus courte,
    on garde la plus longue.
    """
    if len(bumps) <= 1:
        return bumps

    # Trie par longueur décroissante : on traite les plus longues en premier
    sorted_bumps = sorted(bumps, key=lambda b: -b['length_m'])

    kept = []
    for cand in sorted_bumps:
        cand_start = cand['km_start']
        cand_end = cand['km_start'] + cand['length_m'] / 1000
        cand_length_km = cand['length_m'] / 1000

        overlap_too_much = False
        for k in kept:
            k_start = k['km_start']
            k_end = k['km_start'] + k['length_m'] / 1000

            ov_start = max(cand_start, k_start)
            ov_end = min(cand_end, k_end)
            ov_km = max(0, ov_end - ov_start)

            # Si plus de 30% de la plus courte chevauche, on rejette la candidate
            shortest_length_km = min(cand_length_km, k['length_m'] / 1000)
            if shortest_length_km > 0 and ov_km / shortest_length_km > overlap_threshold:
                overlap_too_much = True
                break

        if not overlap_too_much:
            kept.append(cand)

    # Renumérote et retrie par position
    kept.sort(key=lambda b: b['km_start'])
    for i, b in enumerate(kept):
        b['num'] = i + 1

    return kept



# ============================================================================
# VISUALISATION DES PROFILS DE BOSSES
# ============================================================================

def _build_full_profile(coords, elevations):
    """Construit le profil complet (distance cumulative, altitudes) pour affichage.
    Retourne un dict avec dists_km, elevs, et list des indices des bosses détectées.
    """
    if len(coords) < 2 or len(elevations) != len(coords):
        return None

    cum = [0.0]
    for i in range(1, len(coords)):
        cum.append(cum[-1] + haversine_m(coords[i-1], coords[i]))

    return {
        "dists_km": [d / 1000 for d in cum],
        "elevs": elevations,
        "total_km": cum[-1] / 1000,
    }


def _make_overview_chart(profile, bumps):
    """Profil global du parcours avec bosses surlignées en couleur.
    Returns: plotly figure
    """
    import plotly.graph_objects as go

    dists = profile["dists_km"]
    elevs = profile["elevs"]

    fig = go.Figure()

    # Profil de base (gris clair)
    fig.add_trace(go.Scatter(
        x=dists, y=elevs,
        mode="lines",
        line=dict(color="#94a3b8", width=1.5),
        fill="tozeroy",
        fillcolor="rgba(148, 163, 184, 0.2)",
        name="Parcours",
        hovertemplate="km %{x:.1f}<br>alt %{y:.0f}m<extra></extra>",
    ))

    # Pour chaque bosse, surligne la zone en couleur selon pente
    for b in bumps:
        # Trouve les indices correspondants
        km_start = b["km_start"]
        km_end = km_start + b["length_m"] / 1000

        # Sélectionne les points dans la plage
        idx_start = next((i for i, d in enumerate(dists) if d >= km_start), 0)
        idx_end = next((i for i, d in enumerate(dists) if d >= km_end), len(dists) - 1)

        if idx_end <= idx_start:
            continue

        sub_dists = dists[idx_start:idx_end + 1]
        sub_elevs = elevs[idx_start:idx_end + 1]

        # Couleur selon pente moyenne
        grade = b["grade_avg"]
        if grade < 4:
            color = "rgba(255, 193, 7, 0.6)"      # jaune
            line_color = "#f59e0b"
        elif grade < 6:
            color = "rgba(255, 152, 0, 0.6)"      # orange
            line_color = "#f97316"
        elif grade < 9:
            color = "rgba(244, 67, 54, 0.6)"      # rouge
            line_color = "#ef4444"
        else:
            color = "rgba(136, 14, 79, 0.7)"      # violet sombre (HC)
            line_color = "#9333ea"

        fig.add_trace(go.Scatter(
            x=sub_dists, y=sub_elevs,
            mode="lines",
            line=dict(color=line_color, width=2.5),
            fill="tozeroy",
            fillcolor=color,
            name=f"Bosse #{b['num']}",
            hovertemplate=(f"<b>Bosse #{b['num']}</b><br>"
                           f"L={b['length_m']}m · {grade}%<br>"
                           f"km %{{x:.1f}} · alt %{{y:.0f}}m<extra></extra>"),
            showlegend=False,
        ))

        # Marker au sommet de la bosse avec numéro
        peak_idx = max(range(idx_start, idx_end + 1), key=lambda i: elevs[i])
        fig.add_annotation(
            x=dists[peak_idx], y=elevs[peak_idx],
            text=f"<b>{b['num']}</b>",
            showarrow=True, arrowhead=2, arrowsize=1, arrowwidth=1.5,
            arrowcolor=line_color,
            ax=0, ay=-25,
            bgcolor=line_color,
            font=dict(color="white", size=11),
            borderpad=3,
        )

    fig.update_layout(
        title="Profil du parcours · bosses surlignées",
        xaxis_title="Distance (km)",
        yaxis_title="Altitude (m)",
        hovermode="x unified",
        height=350,
        margin=dict(l=50, r=20, t=50, b=50),
        showlegend=False,
        plot_bgcolor="white",
    )
    fig.update_xaxes(gridcolor="#e5e7eb", showgrid=True)
    fig.update_yaxes(gridcolor="#e5e7eb", showgrid=True)

    return fig


def _make_bump_detail_chart(profile, bump):
    """Profil détaillé d'une seule bosse avec coloration par section de pente.
    Applique un smoothing 500m à l'altitude pour éliminer le bruit GPS/baromètre.
    """
    import plotly.graph_objects as go

    dists = profile["dists_km"]
    elevs_raw = profile["elevs"]

    # Smoothing sur 500m pour gommer le bruit GPS (standard industrie)
    # Le step réel dépend de la densité du GPX, on calcule la fenêtre adaptative
    if len(dists) >= 2:
        avg_step_m = (dists[-1] - dists[0]) * 1000 / max(len(dists) - 1, 1)
        window_500m = max(3, int(500 / avg_step_m))
    else:
        window_500m = 3

    # Smoothing par moyenne mobile centrée
    elevs = []
    half = window_500m // 2
    n = len(elevs_raw)
    for i in range(n):
        lo = max(0, i - half)
        hi = min(n, i + half + 1)
        elevs.append(sum(elevs_raw[lo:hi]) / (hi - lo))

    km_start = bump["km_start"]
    km_end = km_start + bump["length_m"] / 1000

    # Ajoute 200m de contexte avant et après
    km_view_start = max(0, km_start - 0.2)
    km_view_end = min(profile["total_km"], km_end + 0.2)

    idx_start = next((i for i, d in enumerate(dists) if d >= km_view_start), 0)
    idx_end = next((i for i, d in enumerate(dists) if d >= km_view_end), len(dists) - 1)

    sub_dists = dists[idx_start:idx_end + 1]
    sub_elevs = elevs[idx_start:idx_end + 1]

    # Calcule pente par segment de ~50m pour colorier
    fig = go.Figure()

    # Trace principal (en gris)
    fig.add_trace(go.Scatter(
        x=sub_dists, y=sub_elevs,
        mode="lines",
        line=dict(color="#cbd5e1", width=1),
        name="Profil",
        hoverinfo="skip",
        showlegend=False,
    ))

    # Coloration par section selon pente locale (fenêtre 250m glissante)
    # 250m = équilibre entre stabilité et précision sur petites bosses
    if len(sub_dists) >= 2:
        avg_step_m = (sub_dists[-1] - sub_dists[0]) * 1000 / max(len(sub_dists) - 1, 1)
        step_window = max(2, int(250 / avg_step_m))
    else:
        step_window = 4

    if len(sub_dists) > step_window:
        for i in range(len(sub_dists) - step_window):
            d1 = sub_dists[i] * 1000
            d2 = sub_dists[i + step_window] * 1000
            if d2 - d1 < 100:
                continue
            grade_local = (sub_elevs[i + step_window] - sub_elevs[i]) / (d2 - d1) * 100

            # Couleur selon pente
            if grade_local < 0:
                color = "#10b981"  # vert (descente)
            elif grade_local < 3:
                color = "#fbbf24"  # jaune
            elif grade_local < 6:
                color = "#fb923c"  # orange
            elif grade_local < 9:
                color = "#ef4444"  # rouge
            else:
                color = "#9333ea"  # violet

            fig.add_trace(go.Scatter(
                x=sub_dists[i:i + step_window + 1],
                y=sub_elevs[i:i + step_window + 1],

                mode="lines",
                line=dict(color=color, width=4),
                hovertemplate=(f"km %{{x:.2f}}<br>alt %{{y:.0f}}m<br>"
                               f"pente locale {grade_local:.1f}%<extra></extra>"),
                showlegend=False,
            ))

    # Ligne verticale début/fin de la bosse
    fig.add_vline(x=km_start, line=dict(color="#64748b", dash="dash", width=1),
                  annotation_text="Début", annotation_position="top left")
    fig.add_vline(x=km_end, line=dict(color="#64748b", dash="dash", width=1),
                  annotation_text="Sommet", annotation_position="top right")

    # Titre avec infos
    title_text = (f"Bosse #{bump['num']} · {bump['length_m']}m à {bump['grade_avg']}% "
                  f"(max {bump['grade_max_100m']}% sur 100m)")

    fig.update_layout(
        title=title_text,
        xaxis_title="Distance (km)",
        yaxis_title="Altitude (m)",
        hovermode="closest",
        height=280,
        margin=dict(l=50, r=20, t=50, b=50),
        showlegend=False,
        plot_bgcolor="white",
    )
    fig.update_xaxes(gridcolor="#e5e7eb", showgrid=True)
    fig.update_yaxes(gridcolor="#e5e7eb", showgrid=True)

    return fig


def render_bumps_visualization(profile, bumps, key_suffix=""):
    """Affiche la visualisation complète des bosses :
    1. Profil global avec bosses surlignées
    2. Détails par bosse dans des expanders
    """
    if not bumps or not profile:
        return

    # Légende colorée
    st.markdown("""
    <div style='display: flex; gap: 15px; flex-wrap: wrap; font-size: 13px; margin-bottom: 10px;'>
        <span><span style='display:inline-block;width:12px;height:12px;background:#fbbf24;border-radius:2px;'></span> < 4%</span>
        <span><span style='display:inline-block;width:12px;height:12px;background:#fb923c;border-radius:2px;'></span> 4-6%</span>
        <span><span style='display:inline-block;width:12px;height:12px;background:#ef4444;border-radius:2px;'></span> 6-9%</span>
        <span><span style='display:inline-block;width:12px;height:12px;background:#9333ea;border-radius:2px;'></span> > 9%</span>
    </div>
    """, unsafe_allow_html=True)

    # 1. Profil global
    fig_overview = _make_overview_chart(profile, bumps)
    st.plotly_chart(fig_overview, use_container_width=True, key=f"overview_{key_suffix}")

    # 2. Détails par bosse
    with st.expander(f"Voir les profils détaillés {len(bumps)} bosses"):
        for b in bumps:
            fig_detail = _make_bump_detail_chart(profile, b)
            st.plotly_chart(fig_detail, use_container_width=True,
                            key=f"detail_{key_suffix}_{b['num']}")
# ============================================================================
# ALTITUDES
# ============================================================================

def get_route_elevations(route_result):
    """Estime un profil d'altitude pour un itinéraire calculé.

    Le graphe contient le D+ par arête mais pas l'altitude absolue.
    Pour V1, on répartit le D+ total proportionnellement à la distance.
    C'est suffisant pour détecter les bosses (profil relatif), mais
    l'altitude absolue est approximative.
    """
    if not route_result:
        return None
    coords = route_result["coords"]
    total_d_plus = route_result.get("total_d_plus", 0)
    total_length = route_result.get("total_length_m", 0)
    if total_d_plus == 0 or total_length == 0:
        return [100.0] * len(coords)

    elevations = [100.0]
    cum_d = 0.0
    for i in range(1, len(coords)):
        seg = haversine_m(coords[i-1], coords[i])
        cum_d += seg
        ratio = cum_d / max(total_length, 1)
        elevations.append(100.0 + ratio * total_d_plus)
    return elevations


def read_gpx_with_elevations(content):
    """Lit un GPX avec altitudes (cas Strava/Garmin/Wahoo)."""
    import xml.etree.ElementTree as ET
    if isinstance(content, bytes):
        content = content.decode('utf-8')
    coords, elevs = [], []
    try:
        root = ET.fromstring(content)
        for elem in root.iter():
            if elem.tag.endswith('trkpt') or elem.tag.endswith('rtept'):
                lat = float(elem.attrib.get('lat', 0))
                lon = float(elem.attrib.get('lon', 0))
                if not (lat and lon):
                    continue
                coords.append((lat, lon))
                ele = 0.0
                for child in elem:
                    if child.tag.endswith('ele'):
                        try:
                            ele = float(child.text)
                        except (ValueError, TypeError):
                            ele = 0.0
                        break
                elevs.append(ele)
    except Exception as e:
        print(f"Erreur GPX : {e}")
    return coords, elevs

# ============================================================================
# EXPORT EXCEL DES BOSSES
# ============================================================================

def export_bumps_to_excel(bumps, route_name="Itinéraire",
                          mass_total_kg=77, mass_bike_kg=7):

    wb = Workbook()
    ws = wb.active
    ws.title = "Bosses"[:31]

    # Couleurs
    C_HEADER = "FF1A3A5C"
    C_SUB1 = "FF2A609B"
    C_SUB2 = "FF2E86AB"
    C_BAND_LIGHT = "FFFFFFFF"
    C_BAND_ALT = "FFEBF2FA"
    C_POINTU = "FFFFF3CD"
    C_MONTEE = "FFF8D7DA"
    C_PLAT = "FFFFF9C4"
    C_DESCENTE = "FFD4EDDA"
    C_T250 = "FFDBEAFE"
    C_T300 = "FFD1FAE5"
    C_T350 = "FFFCE7F3"

    font_title = Font(color="FFFFFFFF", bold=True, size=14, name="Arial")
    font_sub = Font(color="FFFFFFFF", italic=True, name="Arial")
    font_param = Font(color="FFFFFFFF", name="Arial")
    font_header = Font(color="FFFFFFFF", bold=True, name="Arial")
    font_body = Font(name="Arial")

    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Titre
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f" ANALYSE DES BOSSES — {route_name}"
    c.fill = PatternFill("solid", fgColor=C_HEADER)
    c.font = font_title
    c.alignment = align_left
    ws.row_dimensions[1].height = 28

    # Sous-titre 1
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = "Détection des bosses supérieures à 250m et supérieures à 3%"
    c.fill = PatternFill("solid", fgColor=C_SUB1)
    c.font = font_sub
    c.alignment = align_left

    # Sous-titre 2 (paramètres)
    ws.merge_cells("A3:N3")
    c = ws["A3"]
    rider_kg = mass_total_kg - mass_bike_kg
    c.value = (f"Cycliste : {rider_kg} kg  |  Matériel : {mass_bike_kg} kg  |  "
               f"Masse totale : {mass_total_kg} kg  |  Modèle aéro : CdA = 0,30 m²  |  Crr = 0,004")
    c.fill = PatternFill("solid", fgColor=C_SUB2)
    c.font = font_param
    c.alignment = align_left

    # Headers (ligne 5)
    headers = ["N°", "Km début", "Longueur", "Pente moy", "Pente max sur 100m",
               "Section pointue > 100m", "Après le sommet, analyse sur 2,5km",
               "Temps\n250 W", "Temps\n300 W", "Temps\n350 W",
               "D+\n(m)", "Alt. sommet\n(m)", "FIETS", "Cat."]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col_idx)
        c.value = h
        c.fill = PatternFill("solid", fgColor=C_HEADER)
        c.font = font_header
        c.alignment = align_center
    ws.row_dimensions[5].height = 35

    # Largeurs (12 colonnes initiales + FIETS + Cat.)
    widths = [5, 10, 10, 10, 12, 26, 32, 10, 10, 10, 8, 10, 8, 8]
    for i, w in enumerate(widths, start=1):
        # Pour colonnes au-delà de Z, openpyxl utilise AA, AB...
        if i <= 26:
            col_letter = chr(64+i)
        else:
            col_letter = chr(64 + ((i-1)//26)) + chr(64 + ((i-1) % 26) + 1)
        ws.column_dimensions[col_letter].width = w

    # Données
    for row_idx, b in enumerate(bumps, start=6):
        is_alt = (row_idx % 2 == 0)
        band = C_BAND_ALT if is_alt else C_BAND_LIGHT

        values = [
            b["num"],
            f"{b['km_start']:.1f}".replace(".", ","),
            b["length_m"],
            b["grade_avg"],
            b["grade_max_100m"],
            b["section_pointue"],
            b["after_summit"],
            b["time_250w"],
            b["time_300w"],
            b["time_350w"],
            b["d_plus_m"],
            b["alt_summit_m"],
            b.get("fiets", 0),
            b.get("fiets_cat", "—"),
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.font = font_body
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Couleurs conditionnelles
            if col_idx == 6:
                c.fill = PatternFill("solid", fgColor=C_POINTU if val != "Non" else band)
            elif col_idx == 7:
                vs = str(val)
                if "Descente" in vs:
                    fill = C_DESCENTE
                elif "Montée" in vs:
                    fill = C_MONTEE
                elif "Plat" in vs:
                    fill = C_PLAT
                else:
                    fill = band
                c.fill = PatternFill("solid", fgColor=fill)
            elif col_idx == 8:
                c.fill = PatternFill("solid", fgColor=C_T250)
            elif col_idx == 9:
                c.fill = PatternFill("solid", fgColor=C_T300)
            elif col_idx == 10:
                c.fill = PatternFill("solid", fgColor=C_T350)
            else:
                c.fill = PatternFill("solid", fgColor=band)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================================
# SIDEBAR
# ============================================================================

def render_sidebar():
    """Affiche la sidebar et retourne les paramètres de config."""
    with st.sidebar:
        st.header("⚙️ Configuration")

        profile = st.selectbox(
            "Profil",
            options=["club_road", "solo_casual"],
            format_func=lambda x: "🚴 Club road" if x == "club_road" else "🚲 Solo casual",
        )

        alpha = st.slider(
            "Préférence plaisir vs distance",
            min_value=0.0, max_value=2.5, value=1.0, step=0.1,
            help="0 = plus court, 1 = équilibré, 2+ = priorité plaisir",
        )

        st.divider()
        st.subheader("🏔️ Mode grimpeur")
        use_climber = st.checkbox("Activer budget D+", value=False)
        d_plus_target = None
        if use_climber:
            d_plus_target = st.number_input(
                "D+ cible (m)", value=800, min_value=0, max_value=5000, step=100
            )

        st.divider()
        st.subheader("🗺️ Affichage")
        show_club = st.checkbox("Afficher toutes les traces club", value=False)
        show_uploaded = st.checkbox("Afficher trace uploadée", value=True)

        st.divider()
        if st.button("🔄 Réinitialiser waypoints"):
            for k in ["waypoints", "route_result", "bumps_route"]:
                st.session_state[k] = [] if k == "waypoints" else None
            st.rerun()

        st.divider()
        st.subheader("🔍 Recherche d'adresse")
        address = st.text_input("Adresse ou lieu",
            placeholder="Versailles, Bastille, Chevreuse...")
        if st.button("📍 Ajouter ce point", disabled=not address):
            result = geocode_address(address)
            if result:
                lat, lon, display = result
                st.session_state.waypoints.append((lat, lon))
                st.success(f"✅ {display[:60]}")
                st.rerun()
            else:
                st.error("Adresse non trouvée en IdF")

    return {
        "profile": profile,
        "alpha": alpha,
        "use_climber": use_climber,
        "d_plus_target": d_plus_target,
        "show_club": show_club,
        "show_uploaded": show_uploaded,
    }


# ============================================================================
# CARTE
# ============================================================================

def build_map(config, club_traces):
    """Construit l'objet folium.Map avec tous les calques actifs."""
    waypoints = st.session_state.waypoints
    center = [48.80, 2.30]
    if waypoints:
        center = [np.mean([w[0] for w in waypoints]),
                  np.mean([w[1] for w in waypoints])]

    m = folium.Map(location=center, zoom_start=10, tiles="cartodbpositron")

    # Traces club (~149 simplifiées)
    if config["show_club"]:
        simplified = cached_simplified_traces(club_traces)
        fg = folium.FeatureGroup(name=f"Traces club ({len(simplified)})")
        for t in simplified:
            folium.PolyLine(
                t["coords"], color="purple",
                weight=1.5, opacity=0.4, smooth_factor=2,
                tooltip=t["name"][:40],
            ).add_to(fg)
        fg.add_to(m)

    # Trace uploadée
    if config["show_uploaded"] and st.session_state.uploaded_trace:
        folium.PolyLine(
            st.session_state.uploaded_trace,
            color="orange", weight=4, opacity=0.7,
            tooltip="Trace uploadée",
        ).add_to(m)

    # Itinéraire calculé
    if st.session_state.route_result:
        r = st.session_state.route_result
        folium.PolyLine(
            r["coords"], color="red", weight=4, opacity=0.9,
            tooltip=f"Itinéraire ({r['total_length_m']/1000:.1f} km)",
        ).add_to(m)

    # Waypoints
    for i, (lat, lon) in enumerate(waypoints):
        if i == 0:
            icon, label = folium.Icon(color="green", icon="play"), "🚩 Départ"
        elif i == len(waypoints) - 1 and len(waypoints) > 1:
            icon, label = folium.Icon(color="red", icon="stop"), "🏁 Arrivée"
        else:
            icon, label = folium.Icon(color="blue", icon="info-sign"), f"Waypoint {i}"
        folium.Marker([lat, lon], icon=icon, tooltip=label).add_to(m)

    return m


# ============================================================================
# WAYPOINTS + ROUTING
# ============================================================================

def render_waypoints_panel():
    """Affiche la liste des waypoints placés."""
    st.subheader("📍 Waypoints")
    waypoints = st.session_state.waypoints
    if not waypoints:
        st.write("_Clique sur la carte pour placer un point._")
        return

    for i, (lat, lon) in enumerate(waypoints):
        if i == 0:
            emoji = "🚩"
        elif i == len(waypoints) - 1 and len(waypoints) > 1:
            emoji = "🏁"
        else:
            emoji = f"{i}️⃣"
        st.text(f"{emoji} {lat:.4f}, {lon:.4f}")

    # Détection boucle
    if len(waypoints) >= 3:
        d = ((waypoints[0][0] - waypoints[-1][0])**2 +
             (waypoints[0][1] - waypoints[-1][1])**2) ** 0.5
        if d < 0.005:
            st.success("🔁 Boucle détectée")


def trigger_routing(G, config, club_traces):
    """Bouton calculer + lancement du routing + détection bosses."""
    waypoints = st.session_state.waypoints
    can_route = len(waypoints) >= 2

    if st.button("🚀 Calculer l'itinéraire", type="primary", disabled=not can_route):
        with st.spinner("Calcul..."):
            if config["use_climber"]:
                best = None
                best_diff = float('inf')
                for a in [0.5, 1.0, 1.5, 2.0]:
                    r = route_waypoints(G, waypoints, profile=config["profile"], alpha=a)
                    if r:
                        diff = abs(r["total_d_plus"] - config["d_plus_target"])
                        if diff < best_diff:
                            best = r
                            best_diff = diff
                result = best
            else:
                result = route_waypoints(G, waypoints, profile=config["profile"], alpha=config["alpha"])

        if result:
            st.session_state.route_result = result
            st.session_state.comparison = similarity_to_club(result["coords"], club_traces, top_k=5)
            # Bosses sur itinéraire calculé
            elevs = get_route_elevations(result)
            if elevs:
                st.session_state.bumps_route = detect_bumps(result["coords"], elevs)
            st.rerun()
        else:
            st.error("Pas de chemin trouvé entre ces points")


# ============================================================================
# STATS + SIMILARITÉ
# ============================================================================

def render_route_stats(config):
    """Métriques de l'itinéraire calculé + download GPX."""
    if not st.session_state.route_result:
        return
    r = st.session_state.route_result
    st.subheader("📊 Itinéraire calculé")

    c1, c2 = st.columns(2)
    with c1:
        st.metric("Distance", f"{r['total_length_m']/1000:.1f} km")
        st.metric("Score plaisir", f"{r['mean_score']:.2f}")
    with c2:
        st.metric("D+", f"{r['total_d_plus']:.0f} m")
        st.metric("Waypoints", r.get("n_waypoints", 2))

    gpx_content = export_gpx_string(
        r["coords"],
        name=f"Itineraire_{config['profile']}_a{config['alpha']}",
        description=f"{r['total_length_m']/1000:.1f}km, D+{r['total_d_plus']:.0f}m"
    )
    st.download_button(
        "📥 Télécharger GPX",
        data=gpx_content,
        file_name=f"itineraire_{config['profile']}.gpx",
        mime="application/gpx+xml",
    )


def render_similarity():
    """Affiche similarité de l'itinéraire calculé avec les traces club."""
    if not st.session_state.comparison:
        return
    c = st.session_state.comparison
    st.subheader("🎯 Similarité club")
    st.metric("Coverage", f"{c['coverage']*100:.0f}%")
    st.metric("F1 (top 5)", f"{c['global_f1']*100:.0f}%")

    if c["best_matches"]:
        with st.expander("Top 5 traces similaires"):
            for m in c["best_matches"]:
                st.text(f"{m['name'][:30]} — F1 {m['f1']*100:.0f}%")



# ============================================================================
# UPLOAD GPX
# ============================================================================

def render_gpx_upload(club_traces):
    """Upload GPX, comparaison F1 club, détection bosses."""
    st.subheader("📤 Upload trace")
    uploaded = st.file_uploader("Drop un GPX", type=["gpx"])

    if uploaded:
        content = uploaded.read()
        coords, elevs = read_gpx_with_elevations(content)

        if coords:
            st.session_state.uploaded_trace = coords
            st.session_state.uploaded_elevs = elevs
            st.session_state.uploaded_sim = similarity_to_club(coords, club_traces, top_k=5)
            # Détection bosses avec les vraies altitudes du GPX
            if elevs and max(elevs) > 0:
                st.session_state.bumps_uploaded = detect_bumps(coords, elevs)
            else:
                st.session_state.bumps_uploaded = None
            st.success(f"✅ {len(coords)} points")
        else:
            st.error("Aucun point dans le GPX")
            for k in ["uploaded_trace", "uploaded_elevs", "uploaded_sim", "bumps_uploaded"]:
                st.session_state[k] = None

    # Similarité GPX avec TOUTES les traces club
    if st.session_state.uploaded_sim:
        sim = st.session_state.uploaded_sim
        st.markdown("**📊 GPX vs club**")
        col_a, col_b = st.columns(2)
        with col_a:
            st.metric("Coverage", f"{sim['coverage']*100:.0f}%")
        with col_b:
            st.metric("F1 (top 5)", f"{sim['global_f1']*100:.0f}%")

        if sim["best_matches"]:
            with st.expander("Top 5 traces club similaires au GPX"):
                for m in sim["best_matches"]:
                    st.text(f"{m['name'][:35]} — F1 {m['f1']*100:.0f}%")



# ============================================================================
# VISUALISATION DES BOSSES
# ============================================================================

def _style_bump_row(row):
    """Couleurs ligne par ligne dans le DataFrame Streamlit."""
    styles = [""] * len(row)
    cols = list(row.index)

    if "Section pointue" in cols:
        idx = cols.index("Section pointue")
        if row["Section pointue"] != "Non":
            styles[idx] = "background-color: #FFF3CD"

    if "Après sommet" in cols:
        idx = cols.index("Après sommet")
        val = str(row["Après sommet"])
        if "Descente" in val:
            styles[idx] = "background-color: #D4EDDA"
        elif "Montée" in val:
            styles[idx] = "background-color: #F8D7DA"
        elif "Plat" in val:
            styles[idx] = "background-color: #FFF9C4"

    # Colonnes temps
    for col_name, color in [("Temps 250W", "#DBEAFE"),
                             ("Temps 300W", "#D1FAE5"),
                             ("Temps 350W", "#FCE7F3")]:
        if col_name in cols:
            idx = cols.index(col_name)
            if not styles[idx]:
                styles[idx] = f"background-color: {color}"

    return styles


def _show_bumps_table(bumps, name, fname, profile=None):
    """Affiche profil interactif + tableau bosses + bouton Excel."""
    if not bumps:
        st.info("Aucune bosse détectée (seuils : >250m et >3%)")
        return

    # Visualisation interactive en premier (si profil disponible)
    if profile:
        render_bumps_visualization(profile, bumps, key_suffix=fname)
        st.divider()

    df = pd.DataFrame([{
        "N°": b["num"],
        "Km début": f"{b['km_start']:.1f}".replace(".", ","),
        "Longueur (m)": b["length_m"],
        "Pente moy (%)": b["grade_avg"],
        "Pente max 100m": b["grade_max_100m"],
        "Section pointue": b["section_pointue"],
        "Après sommet": b["after_summit"],
        "Temps 250W": b["time_250w"],
        "Temps 300W": b["time_300w"],
        "Temps 350W": b["time_350w"],
        "D+ (m)": b["d_plus_m"],
        "Alt. sommet": b["alt_summit_m"],
        "FIETS": b.get("fiets", "—"),
        "Cat.": b.get("fiets_cat", "—"),
    } for b in bumps])

    styled = df.style.apply(_style_bump_row, axis=1)
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Pente moy (%)": st.column_config.NumberColumn(format="%.1f"),
            "Pente max 100m": st.column_config.NumberColumn(format="%.1f"),
        },
    )

    # Bouton Excel stylé
    xlsx_bytes = export_bumps_to_excel(bumps, route_name=name)
    st.download_button(
        f"📥 Télécharger Excel ({name})",
        data=xlsx_bytes,
        file_name=f"bosses_{fname}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{fname}",
    )

    # Stats
    total_dp = sum(b["d_plus_m"] for b in bumps)
    avg_g = np.mean([b['grade_avg'] for b in bumps])
    st.caption(f"📊 {len(bumps)} bosses · {total_dp}m D+ cumulé · "
               f"pente moyenne {avg_g:.1f}%")


def render_bumps_section():
    """Section bosses sous la carte. Tabs si itinéraire + GPX."""
    bumps_route = st.session_state.get("bumps_route")
    bumps_uploaded = st.session_state.get("bumps_uploaded")

    if not bumps_route and not bumps_uploaded:
        return

    st.divider()
    st.subheader("⛰️ Analyse des bosses")
    st.caption("Détection : longueur ≥ 250m ET pente moyenne ≥ 3% — "
               "Simulation : 77 kg, CdA 0.30 m², Crr 0.004")

    to_show = []
    if bumps_route:
        # Construit le profil de l'itinéraire calculé
        r = st.session_state.route_result
        if r:
            elevs = get_route_elevations(r)
            profile = _build_full_profile(r["coords"], elevs) if elevs else None
        else:
            profile = None
        to_show.append(("Itinéraire calculé", bumps_route, "itineraire", profile))

    if bumps_uploaded:
        # Construit le profil du GPX uploadé
        coords = st.session_state.uploaded_trace
        elevs = st.session_state.uploaded_elevs
        if coords and elevs:
            profile = _build_full_profile(coords, elevs)
        else:
            profile = None
        to_show.append(("GPX uploadé", bumps_uploaded, "gpx_upload", profile))

    if len(to_show) == 1:
        name, bumps, fname, profile = to_show[0]
        _show_bumps_table(bumps, name, fname, profile)
    else:
        tabs = st.tabs([t[0] for t in to_show])
        for tab, (name, bumps, fname, profile) in zip(tabs, to_show):
            with tab:
                _show_bumps_table(bumps, name, fname, profile)



# ============================================================================
# MAIN
# ============================================================================

def main():
    init_session_state()

    G = cached_graph()
    club_traces = cached_traces()

    config = render_sidebar()

    col_map, col_info = st.columns([3, 1])

    with col_map:
        st.subheader("Carte")

        n_wp = len(st.session_state.waypoints)
        if n_wp == 0:
            st.info("👇 Clique sur la carte pour placer ton **départ**")
        elif n_wp == 1:
            st.info("👇 Clique pour placer ton **arrivée** ou un waypoint")
        else:
            st.info(f"✅ {n_wp} points · ajoute d'autres waypoints ou clique **Calculer**")

        m = build_map(config, club_traces)
        map_data = st_folium(m, height=600, width=None, returned_objects=["last_clicked"])

        if map_data and map_data.get("last_clicked"):
            clicked = map_data["last_clicked"]
            new_wp = (clicked["lat"], clicked["lng"])
            if not st.session_state.waypoints or st.session_state.waypoints[-1] != new_wp:
                st.session_state.waypoints.append(new_wp)
                st.rerun()

    with col_info:
        render_waypoints_panel()
        st.divider()
        trigger_routing(G, config, club_traces)
        st.divider()
        render_route_stats(config)
        st.divider()
        render_similarity()
        st.divider()
        render_gpx_upload(club_traces)

    # Section bosses en pleine largeur sous la carte
    render_bumps_section()

    # Footer
    st.divider()
    with st.expander("ℹ️ Aide"):
        st.markdown("""
        """)


if __name__ == "__main__":
    main()
